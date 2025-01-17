import tkinter as tk
from tkinter import filedialog, messagebox
import csv
from openpyxl import load_workbook
import re


csv_entries = []
OLD_LENGTH = 0

# Define file paths for automatic selection
RAN1_PATH = "C:/Users/Omart/Desktop/GUC/Orange Internship docs/Nokia/RAN1/MR.conf"  # Replace with actual path for RAN1
RAN2_PATH = "C:/Users/Omart/Desktop/GUC/Orange Internship docs/Nokia/RAN2/MR.conf"  # Replace with actual path for RAN2

def upload_csv_file():
    global csv_entries
    csv_file_path = filedialog.askopenfilename(
        title="Select CSV or Excel File",
        filetypes=[("CSV or Excel files", "*.csv;*.xlsx")]
    )
    if not csv_file_path:
        messagebox.showinfo("File Selection", "No file selected.")
        return

    try:
        if csv_file_path.endswith(".csv"):
            # Read the CSV file
            with open(csv_file_path, 'r') as csv_file:
                reader = csv.reader(csv_file)
                next(reader, None)  # Skip the header row, if any
                csv_entries = [row[0].strip() for row in reader if row and row[0].strip()]

        elif csv_file_path.endswith(".xlsx"):
            # Read the Excel file using openpyxl
            workbook = load_workbook(filename=csv_file_path)
            sheet = workbook.active  # Get the active sheet

            # Extract entries from the first column, skipping the header row
            csv_entries = [
                str(row[0]).strip()
                for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True)
                if row[0]  # Skip empty rows
            ]

        else:
            messagebox.showerror("Error", "Unsupported file type. Please upload a .csv or .xlsx file.")
            return

        # Populate the Text widget with the loaded entries
        entry_text.delete("1.0", "end")  # Clear existing text
        entry_text.insert("1.0", "\n".join(csv_entries))  # Insert entries

        messagebox.showinfo("Success", f"{len(csv_entries)} entries loaded from the file.")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to read file: {e}")

# Function to extract the site code from a line
#def get_site_code(site_line):
    #if '_' in site_line:
        #return site_line.rsplit('_', 1)[-1]
    #return None

def get_site_code(site_line):
    # Regular expression to find any combination of 4 digits followed by 2 letters
    match = re.search(r'\d{4}[A-Za-z]{2}', site_line)
    if match:
        return match.group()  # Return the matched site code
    return None

# Function to remove duplicates and return removed lines
def remove_duplicates_and_count(conf_file_path, new_sites):
    try:
        with open(conf_file_path, 'r') as conf_file:
            lines = conf_file.readlines()

        site_code_set = set(get_site_code(site) for site in new_sites if get_site_code(site))
        updated_lines = []
        duplicates_removed = []

        for line in lines:
            stripped_line = line.strip()
            site_code = get_site_code(stripped_line)
            if site_code and site_code in site_code_set:
                duplicates_removed.append(stripped_line)
            else:
                updated_lines.append(line)

        with open(conf_file_path, 'w') as conf_file:
            conf_file.writelines(updated_lines)

        return duplicates_removed, len(duplicates_removed), len(updated_lines)
    except FileNotFoundError:
        messagebox.showerror("Error", "The selected file does not exist.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")
        return [], 0, 0

# Function to insert text, count duplicates, and remove them
def insert_text_to_conf():
    site_entries = entry_text.get("1.0", "end").strip().splitlines()
    valid_entries = []
    invalid_entries = []

    for entry in site_entries:
        try:
            validate_entry(entry)
            valid_entries.append(entry)
        except ValueError as e:
            invalid_entries.append(f"{entry}: {str(e)}")

    if not site_entries or all(not line.strip() for line in site_entries):
        messagebox.showerror("Error", "No text to insert.")
        return

    valid_entries = list(set(valid_entries))
    OLD_LENGTH = len(valid_entries)

    ran1_entries = [entry for entry in valid_entries if get_code_and_file(entry)[1] == "RAN1"]
    ran2_entries = [entry for entry in valid_entries if get_code_and_file(entry)[1] == "RAN2"]


    def process_entries(conf_path, entries):
        duplicates_removed, lines_deleted, total_lines_after = remove_duplicates_and_count(conf_path, entries)
        with open(conf_path, 'a') as conf_file:
            for site_entry in entries:
                conf_file.write(get_code_and_file(site_entry)[0] + "," + site_entry.upper() + "\n")
        return duplicates_removed, lines_deleted, total_lines_after

    duplicates1, lines_deleted1, total1 = process_entries(RAN1_PATH, ran1_entries)
    duplicates2, lines_deleted2, total2 = process_entries(RAN2_PATH, ran2_entries)

    result_display.config(state="normal")
    result_display.delete("1.0", "end")

    # Display invalid entries first
    if invalid_entries:
        seen = set()
        invalid_entries = [entry for entry in invalid_entries if not (entry in seen or seen.add(entry))]
        result_display.insert("1.0", f"Validation Errors (Count: {len(invalid_entries)}):\n")
        result_display.insert("end", "\n".join(invalid_entries) + "\n\n")

    # Now display the duplicates and insertion results
    result_display.insert("end", f"RAN1 File:\n")
    result_display.insert("end", f"Removed Duplicates (Count: {lines_deleted1}):\n")
    result_display.insert("end", "\n".join(duplicates1) + "\n\n")
    result_display.insert("end", f"Total Lines After Insertion: {total1}\n\n")

    result_display.insert("end", f"RAN2 File:\n")
    result_display.insert("end", f"Removed Duplicates (Count: {lines_deleted2}):\n")
    result_display.insert("end", "\n".join(duplicates2) + "\n\n")
    result_display.insert("end", f"Total Lines After Insertion: {total2}\n")

    messagebox.showinfo("Success", f"Entries successfully inserted: {len(valid_entries)}\nInvalid entries: {len(invalid_entries)}")

    result_display.config(state="disabled")

# Validation
def validate_entry(entry):
    check_format(entry)
    res = get_code_and_file(entry)
    print(res)
    return res[0] # returns pre comma part if valid, otherwise an error would've been thrown

def get_code_and_file(entry): # checks for existance of site code and returns pre comma if valid site code
    # Split the entry
    parts = entry.split("_")
    if(len(parts)<1):
        raise ValueError("Incorrect format: site name is not following any of the correct naming conventions")
    #code = parts[len(parts)-1]
    code = get_site_code(entry)
    if not code[0].isdigit():
        raise ValueError("Invalid site code: The code must start with a number.")
    site = ""
    file = ""
    last_two_chars = code[-2:].upper()
    if last_two_chars == "AL":
        site = "Alex"
        file = "RAN2"
    elif last_two_chars == "DE":
        site = "SYS_DELTA_NORTH"
        file = "RAN2"
    elif last_two_chars == "UP" or last_two_chars == "SI":
        site = "HUA_MBV_NLG"
        file = "RAN1"
    else:
         raise ValueError("Incorrect Site Code Or Site Code Does Not Exist: The site code must end with one of the following: AL, DE, SI, or UP.")
    
    return [site, file]

# print(getCodeAndFile("m_npAO"))

def check_format(entry): # checks format: no spaces and number of fields of the entry is not less than 4
    parts = entry.split("_")
    # if len(parts) < 4:
    #     raise ValueError("Incorrect format: Missing one or more fields (must have at least 4 parts).")
    # elif len(parts) > 5:
    #     raise ValueError("Incorrect format: Extra fields entered (must not exceed 5 parts).")
    for part in parts:
        if " " in part:
            raise ValueError("Incorrect format: Fields must not contain spaces.")
# Validation End
def on_enter(e, button):
    button.config(bg="#ff9f00")  # Change the background color on hover

def on_leave(e, button):
    button.config(bg="#ff7900")  # Revert the background color when the mouse leaves

# Create the main window
root = tk.Tk()
root.title("Orange Site Entry Tool")
root.configure(bg="#1a1a1a")  # Set background to dark color
root.state('zoomed')  # Fullscreen mode

# Centering and styles
frame = tk.Frame(root, bg="#1a1a1a")
frame.pack(expand=True)

# Add a bold title with white font
title_label = tk.Label(frame, text="Site Entry Configuration Tool", fg="white", bg="#1a1a1a", font=("Arial", 24, "bold"))
title_label.pack(pady=20)  # Add some padding at the top

# Button and text styles
button_style = {"bg": "#ff7900", "fg": "white", "font": ("Arial", 14, "bold"), "relief": "flat", "highlightthickness": 0}
label_style = {"bg": "#1a1a1a", "fg": "white", "font": ("Arial", 12)}
text_style = {
    "bg": "#2b2b2b", "fg": "white", "insertbackground": "white", "font": ("Consolas", 11),
    "bd": 0, "highlightthickness": 0, "relief": "flat"
}

# Logo Placeholder
# Load the logo image (adjust the path to your logo)
logo_image = tk.PhotoImage(file='C:/Users/Omart/Desktop/GUC/Orange Internship docs/WhatsApp Image 2025-01-17 at 10.16.39 PM.png')

# Logo Placeholder
logo_frame = tk.Frame(root, bg="#1a1a1a")
logo_frame.place(relx=0, rely=0, anchor='nw', x=20, y=20)  # Position near top left with some padding
logo_label = tk.Label(logo_frame, image=logo_image, bg="#1a1a1a")
logo_label.pack()

# UI Elements with hover effects
upload_button = tk.Button(frame, text="Upload CSV File", command=upload_csv_file, **button_style)
upload_button.pack(pady=10)
upload_button.bind("<Enter>", lambda e: on_enter(e, upload_button))
upload_button.bind("<Leave>", lambda e: on_leave(e, upload_button))

tk.Label(frame, text="Text to Insert (one entry per line):", **label_style).pack(pady=5)
entry_text = tk.Text(frame, height=10, width=60, **text_style)
entry_text.pack(pady=5)

insert_button = tk.Button(frame, text="Insert Text and Remove Duplicates", command=insert_text_to_conf, **button_style)
insert_button.pack(pady=(15, 60))  # Increased bottom padding
insert_button.bind("<Enter>", lambda e: on_enter(e, insert_button))
insert_button.bind("<Leave>", lambda e: on_leave(e, insert_button))

tk.Label(frame, text="Result Information:", **label_style).pack(pady=5)

result_display = tk.Text(frame, height=15, width=90, **text_style)
result_display.config(state="disabled")
result_display.pack(pady=5)

# Start the GUI loop
root.mainloop()